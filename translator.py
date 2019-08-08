#translator.py
from openpyxl import load_workbook
from openpyxl import Workbook
from googletrans import Translator
import googletrans

##########CONFIG YOUR EXCEL###########
vocabfile = 'allword.xlsx' #put your vocab in Column [A]
transfile = 'result_translate.xlsx' #change file name
lang_dest = 'zh-cn' #translate to
sheetname = 'Sheet1' #sheetname
totalvocab = 3
######################################

#chinese: 'zh-cn'
#german: de
#japanese: ja

#uncomment below line for see languages code
#print(googletrans.LANGUAGES)
try:
	excelword = load_workbook(vocabfile)

	row = excelword[sheetname]

	allwords = []
	alltrans = []

	for i in range(1,totalvocab + 1):
		t = row.cell(row=i,column=1).value
		allwords.append(t)

	def exporttoexcel(translist):
		excelfile = Workbook()
		row = excelfile.active

		for trs in translist:
			row.append(trs)

		excelfile.save(transfile)
		print('Done')

	#show languages code

	LAM = Translator()

	def Trans(wd,dt='en'):
		word = LAM.translate(wd,dest=dt)
		print('Word: {} Trans: {}'.format(wd,word.text))
		print('Pron: {}'.format(word.pronunciation))
		print('------')
		return [wd,word.text,word.pronunciation]

	#allword = ['สวัสดี','สบายดีไหม','รถยนต์']

	for w in allwords:
		t = Trans(w,lang_dest)
		alltrans.append(t)

	exporttoexcel(alltrans)
except:
	print('please input filename on the top')